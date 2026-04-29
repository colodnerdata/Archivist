import base64
import io
import logging

logger = logging.getLogger(__name__)

_TEXT_EXTENSIONS = {
    ".txt", ".md", ".py", ".js", ".ts", ".jsx", ".tsx", ".css", ".html",
    ".htm", ".xml", ".json", ".yaml", ".yml", ".toml", ".cfg", ".conf",
    ".sh", ".bat", ".ps1", ".rb", ".go", ".java", ".c", ".cpp", ".h",
    ".hpp", ".cs", ".rs", ".php", ".sql", ".r", ".csv", ".tsv",
}

_IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".raw",
    ".psd", ".webp", ".heic", ".heif",
}


def extract(file_path: str, extension: str, config: dict) -> tuple[str, str]:
    """
    Returns (content_type, content).
    content_type: 'text' | 'complex_pdf' | 'image_b64' | 'error'

    'complex_pdf' signals that the PDF contains embedded images, very sparse
    text, or other indicators that it is more than plain text.  The caller
    should route it to a vision/thinking model rather than a plain text model.
    """
    ext = extension.lower()
    try:
        if ext == ".docx":
            return ("text", _extract_docx(file_path))
        elif ext == ".pdf":
            text, is_complex = _extract_pdf(file_path)
            return ("complex_pdf" if is_complex else "text", text)
        elif ext == ".xlsx":
            return ("text", _extract_xlsx(file_path))
        elif ext in _IMAGE_EXTENSIONS:
            return ("image_b64", _extract_image(file_path))
        elif ext in _TEXT_EXTENSIONS:
            return ("text", _extract_text(file_path))
        else:
            # Try reading as text, fall back to error
            try:
                return ("text", _extract_text(file_path))
            except Exception:
                return ("error", f"Unknown extension {ext!r} and not readable as text")
    except Exception as e:
        return ("error", str(e))


def _extract_docx(file_path: str) -> str:
    from docx import Document
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


# Average chars-per-page below this threshold indicates sparse text content
# (scanned pages, chart-heavy slides, forms, etc.) and triggers the thinking model.
_PDF_SPARSE_TEXT_THRESHOLD = 100


def _extract_pdf(file_path: str) -> tuple[str, bool]:
    """Return (extracted_text, is_complex).

    is_complex is True when the PDF has embedded images on any page, or when
    the average extracted character count per page is below
    _PDF_SPARSE_TEXT_THRESHOLD, suggesting the document is scanned, form-heavy,
    or relies on charts/diagrams rather than prose.
    """
    import pdfplumber
    pages: list[str] = []
    total_chars = 0
    has_embedded_images = False

    with pdfplumber.open(file_path) as pdf:
        n_pages = len(pdf.pages)
        for page in pdf.pages:
            if page.images:
                has_embedded_images = True
            text = page.extract_text()
            if text:
                total_chars += len(text)
                pages.append(text)

    sparse = n_pages > 0 and (total_chars / n_pages) < _PDF_SPARSE_TEXT_THRESHOLD
    is_complex = has_embedded_images or sparse
    return "\n".join(pages), is_complex


def _extract_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"Sheet: {sheet_name}")
        rows_seen = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                parts.append("\t".join(cells))
                rows_seen += 1
                if rows_seen >= 20:
                    break
    wb.close()
    return "\n".join(parts)


def _extract_text(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _extract_image(file_path: str) -> str:
    from PIL import Image
    with Image.open(file_path) as img:
        img = img.convert("RGB")
        img.thumbnail((1024, 1024), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        return base64.b64encode(buf.getvalue()).decode("utf-8")


def _truncate_to_tokens(text: str, max_tokens: int) -> str:
    words = text.split()
    if len(words) <= max_tokens:
        return text
    return " ".join(words[:max_tokens])
